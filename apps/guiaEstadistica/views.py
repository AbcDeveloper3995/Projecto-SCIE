import json

from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, CreateView, TemplateView, UpdateView
from utils import getCuestionarios
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

from apps.guiaEstadistica.forms import *
from apps.guiaEstadistica.models import guiaEstadistica, cuestionario
from apps.indicadores.forms import seccion, clasificadorIndicadores
from apps.indicadores.models import Indicadores
from apps.seccion.forms import nomencladorColumna, instanciaSeccion, instanciaForm, verificacionForm
from apps.seccion.models import clasificadorPeriodo, verificacion


class listarGuiasView(ListView):
    template_name = 'guiaEstadistica/listarGuia.html'
    model = guiaEstadistica
    context_object_name = 'guias'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Listado de Guias'
        return context

class crearGuiasView(CreateView):
    template_name = 'guiaEstadistica/crearGuia.html'
    model = guiaEstadistica
    form_class = guiaEstadisticaForm
    success_url = reverse_lazy('guia:listarGuias')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Creacion de guia'
        context['action'] = 'add'
        context['guias'] = self.getGuias()
        return context

    def getGuias(self):
        query = guiaEstadistica.objects.all()
        return query

class updateGuiaView(UpdateView):
    model = guiaEstadistica
    form_class = guiaEstadisticaForm
    template_name = 'guiaEstadistica/crearGuia.html'
    success_url = reverse_lazy('guia:listarGuias')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Edicion de guia'
        return context

class eliminarGuia(TemplateView):

    def get(self, request, *args, **kwargs):
        guia = get_object_or_404(guiaEstadistica, id=self.kwargs['pk'])
        guia.delete()
        messages.success(self.request,"La guia " + guia.nombre + " ha sido eliminada correctamente.")
        return redirect('guia:listarGuias')

class captarDatosView(TemplateView):
    template_name = 'guiaEstadistica/captarDatos.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        print(request.POST)
        action = request.POST['action']
        lastCuestionario = cuestionario.objects.last()
        query = instanciaSeccion.objects.filter(seccion_id=request.POST['id_seccion'], cuestionario_fk_id=lastCuestionario.id)
        try:
           if action == 'mostrarInstancias' and query.exists():
               data = []
               for i in query:
                    data.append(i.toJSON())
           else:
               data['error'] = 'error'
        except Exception as e:
            data['error'] =str(e)
        return JsonResponse(data, safe= False)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['guia'] = self.getGuia()
        if context['guia'] == None:
            redirect('usuario:home')
        else:
            context['secciones'] = self.getSecciones()
            context['verificacionForm'] = verificacionForm
            context['columnas'] = self.getCol()
            context['instancias'] = self.getInstancias()
            context['instanciaForm'] = instanciaForm
            context['datos'] = self.getDatos()
            print(context['datos'])
            context['universo'] = self.getUniverso()
            context['titulo'] = 'Informacion vinculada a la entidad:'
        return context

    def getGuia(self):
        try:
            guia = guiaEstadistica.objects.get(activo=True)
            return guia
        except Exception as e:
            messages.error(self.request, "No hay guias activas en este momento.")
            return None

    def getSecciones(self):
        data = []
        guia = self.getGuia()
        secciones = seccion.objects.filter(guia_id=guia.id)
        for i in secciones:
            data.append(i.id)
        return data

    def getDatos(self):

        guia = self.getGuia()
        data = {}
        secciones = seccion.objects.filter(guia_id=guia.id)
        for i in secciones:
            data[i.nombre] = self.getGrupoIndicador(i.id)
            data.copy()
        return data

    def getGrupoIndicador(self, idSeccion):
        aux = []
        if clasificadorIndicadores.objects.filter(seccion_id=idSeccion).exists():
            grupoInd = clasificadorIndicadores.objects.filter(seccion_id=idSeccion)
            for i in grupoInd:
                aux.append(i)
                self.getIndicador(i, aux)
            return aux
        else:
            secciones = self.getSecciones()
            for i in secciones:
                if i == idSeccion:
                    query = seccion.objects.get(id=i)
                    aux.append(query)
            return aux

    def getIndicador(self, i, aux):
        indicadores = Indicadores.objects.filter(clasificadorIndicadores_id=i.id).order_by('fechaCreacion')
        for i in indicadores:
            aux.append(i)

    def getCol(self):
        data ={}
        secciones = self.getSecciones()
        for i in secciones:
            col = nomencladorColumna.objects.filter(seccion_id=i)
            data[i] = col
            data.copy()
        return data

    def getInstancias(self):
        data = {}
        secciones = self.getSecciones()
        for i in secciones:
            obj = instanciaSeccion.objects.filter(seccion_id=i)
            data[i] = obj
            data.copy()
        return data

    def getUniverso(self):
        data = []
        if self.request.user.is_superuser:
            dataUniverso = universoEntidades.objects.all()
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.pinar'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=21)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_artemisa'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=22)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.habana'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=23)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_mayabeque'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=24)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_matanzas'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=25)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_villa_clara'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=26)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_cienfuegos'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=27)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_santi_spiritu'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=28)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_ciego'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=29)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_camaguey'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=30)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_las_tunas'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=31)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_holguin'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=32)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_granma'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=33)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_santiago'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=34)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_guantanamo'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=35)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_la_isla'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=40)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_ZED_mariel'):
            dataUniverso = universoEntidades.objects.filter(guia__activo=True).filter(entidad_codigo__ote_codigo__codigo__exact=41)
            for i in dataUniverso:
                data.append(i.entidad_codigo)
        return data

class crearUniversoView(CreateView):
    template_name = 'entidad/crearUniverso.html'
    model = universoEntidades
    form_class = universoForm
    success_url = reverse_lazy('guia:listarUniverso')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Creacion de universo'
        return context

class updateUniversoView(UpdateView):
    model = universoEntidades
    form_class = universoForm
    template_name = 'entidad/crearUniverso.html'
    success_url = reverse_lazy('guia:listarUniverso')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Edicion de universo'
        return context

class listarUniversoView(ListView):
    model = universoEntidades
    template_name = 'entidad/listarUniverso.html'
    context_object_name = 'universo'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Universos definidos'
        return context

class eliminarUniverso(TemplateView):

    def get(self, request, *args, **kwargs):
        universo = get_object_or_404(universoEntidades, id=self.kwargs['pk'])
        universo.delete()
        messages.success(self.request, "La entidad " + universo.entidad_codigo.nombre_CI + " ha sido eliminada del universo correctamente.")
        return redirect('guia:listarUniverso')

class dataUniversoView(TemplateView):
    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        print(request.POST)
        dataJson = json.loads(request.POST['data'])
        enitdad = Entidad.objects.all()
        try:
            for j in enitdad:
                 for i in dataJson:
                     if i == j.codigo_CI:
                         print(i,j , 'Coinciden')
                         self.crearUniverso(j)
            data['exito'] = 'Universo creado correctamente.'
        except Exception as e:
            data['error'] =str(e)
        return JsonResponse(data, safe= False)


    def getGuia(self):
        guia = guiaEstadistica.objects.get(activo=True)
        return guia

    def crearUniverso(self, Entidad):
            obj = universoEntidades(
                guia=self.getGuia(),
                entidad_codigo=Entidad
            )
            obj.save()

class dataCaptacion(captarDatosView):
    template_name = 'guiaEstadistica/captarDatos.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        campos = dict(request.POST)
        action = campos.get('action')
        entidad = campos.get('Entidad')
        try:
            if action[0] == 'dataCaptacion':
                objCuestionario = self.createCuestionario(entidad)
                if objCuestionario == False:
                    data['error'] = 'La entidad seleccionada ya ha sido controlada.'
                else:
                    for clave, valor in campos.items():
                        if clave != 'action':
                            queryPreEval = PreguntasEvaluadas(
                                captacion_id=objCuestionario,
                                pregunta=clave,
                                respuesta=valor[0]
                            )
                            queryPreEval.save()
                    data['exito'] = 'Informacion guardada correctamente.'
            if action[0] == 'crearInstancia':
                objSeccion = seccion.objects.get(id=request.POST['seccion_id'])
                lastCuestionario = self.getLastCuestionario()
                if objSeccion.periodo_id.tipo == "Anual":
                    instancia = instanciaSeccion(
                        seccion_id_id=request.POST['seccion_id'],
                        cuestionario_fk=lastCuestionario,
                        codigo_id_id=request.POST['codigo_id'],
                        columna_id_id=request.POST['columna_id'],
                        registro_1=request.POST['registro_1'],
                        registro_2=None,
                        registro_3=None,
                        modelo_1=request.POST['modelo_1'],
                        modelo_2=None,
                        modelo_3=None,
                    )
                    instancia.save()
                else:
                    instancia = instanciaSeccion(
                        seccion_id_id=request.POST['seccion_id'],
                        cuestionario_fk=lastCuestionario,
                        codigo_id_id=request.POST['codigo_id'],
                        columna_id_id=request.POST['columna_id'],
                        registro_1=request.POST['registro_1'],
                        registro_2=request.POST['registro_2'],
                        registro_3=request.POST['registro_3'],
                        modelo_1=request.POST['modelo_1'],
                        modelo_2=request.POST['modelo_2'],
                        modelo_3=request.POST['modelo_3'],
                    )
                    instancia.save()
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def createCuestionario(self, entidad):
        try:
            query = Entidad.objects.get(nombre_CI__exact=entidad[0])
            print(query)
            objCuestionario = cuestionario(
                    guia=self.getGuia(),
                    entidad_codigo=query
                )
            objCuestionario.save()
            return objCuestionario
        except:
            return False


    def getLastCuestionario(self):
        query = cuestionario.objects.last()
        return query

class guiaCaptada(ListView):
    template_name = 'guiaEstadistica/guiaCaptada.html'
    model = cuestionario
    context_object_name = 'cuestionarios'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.is_superuser:
            context['cuestionarios'] = cuestionario.objects.all()
        elif self.request.user.has_perm('guiaEstadistica.pinar'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=21)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_artemisa'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=22)
        elif self.request.user.has_perm('guiaEstadistica.habana'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=23)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_mayabeque'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=24)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_matanzas'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=25)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_villa_clara'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=26)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_cienfuegos'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=27)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_santi_spiritu'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=28)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_ciego'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=29)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_camaguey'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=30)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_las_tunas'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=31)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_holguin'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=32)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_granma'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=33)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_santiago'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=34)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_guantanamo'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=35)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_la_isla'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=40)
        elif self.request.user.has_perm('guiaEstadistica.ver_entidades_ZED_mariel'):
            context['cuestionarios'] = cuestionario.objects.filter(guia__activo=True).filter(
                entidad_codigo__ote_codigo__codigo__exact=41)
        context['titulo'] = 'Cuestionarios captados'
        context['titulo2'] = 'Informacion Captada'
        context['titulo3'] = 'Modificar Preguntas'
        context['titulo4'] = 'Modificar Instancias'
        return context

class eliminarGuiaCaptada(TemplateView):

    def get(self, request, *args, **kwargs):
        guia = get_object_or_404(cuestionario, id=self.kwargs['pk'])
        guia.delete()
        messages.success(self.request,"El cuestionario captado a " + guia.entidad_codigo.nombre_CI + " ha sido eliminado correctamente.")
        return redirect('guia:guiaCaptada')

class informacionCaptada(TemplateView):

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'informacionCaptada':
                data = []
                preguntasCaptadas = PreguntasEvaluadas.objects.filter(captacion_id=request.POST['id'])
                for i in preguntasCaptadas:
                    data.append(i.toJSON())
                instancias = instanciaSeccion.objects.filter(cuestionario_fk=request.POST['id'])
                for i in instancias:
                    data.append(i.toJSON())
            else:
                pass
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

class seccionCaptada(TemplateView):

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'seccionCaptada':
                data = []
                query_seccion = seccion.objects.get(id=request.POST['id'])
                if query_seccion.periodo_id =='Anual':
                    instancia = instanciaSeccion.objects.filter(seccion_id_id=query_seccion.id)
                    for i in instancia:
                        data.append(i.toJSON())
                else:
                    instancia = instanciaSeccion.objects.filter(seccion_id_id=query_seccion.id)
                    for i in instancia:
                        data.append(i.toJSON())
            else:
                pass
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

class modificarPreguntasView(TemplateView):

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        campos = dict(request.POST)
        print(campos)
        try:
            if action == 'cargarPreguntas':
               data = []
               query = PreguntasEvaluadas.objects.filter(captacion_id=request.POST['id'])
               for i in query:
                   data.append(i.toJSON())
            if action == 'modificarPreguntas':
                for clave, valor in campos.items():
                    if clave != 'action' and clave != 'id':
                        query = PreguntasEvaluadas.objects.filter(pregunta=clave).filter(captacion_id=campos['id'][0])
                        pregunta = query[0]
                        pregunta.respuesta = valor[0]
                        pregunta.save()

        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)


class reporteGeneralExcel(TemplateView):
    # def get(self, request, *args, **kwargs):
    #     query_cuestionario = self.getCuestionarios()
    #     wb = Workbook()
    #     ws = wb.active
    #     self.estilosCelda(ws['A1'])
    #     ws['A1'] = 'Identificacion'
    #     ws.row_dimensions[1].height = 25
    #     ws.column_dimensions['A'].width = 20
    #     ws.merge_cells('A1:C1')
    #     ws.merge_cells('A2:A3')
    #     ws.merge_cells('B2:B3')
    #     ws.merge_cells('C2:C3')
    #     ws['A2'] = 'Empresa'
    #     ws['B2'] = 'Codigo'
    #     ws['C2'] = 'DPA'
    #
    #     self.estilosCelda(ws['D1'])
    #     ws['D1'] = 'Aspectos Generales'
    #     ws.merge_cells('D1:H1')
    #     ws.column_dimensions['D'].width = 20
    #     ws['D2'] = 'Certificado REEUP'
    #     ws.merge_cells('D2:D3')
    #     ws.column_dimensions['E'].width = 20
    #     ws['E2'] = 'Fecha Certificado'
    #     ws.merge_cells('E2:E3')
    #     ws.column_dimensions['F'].width = 20
    #     ws['F2'] = 'Ubicacion visible'
    #     ws.merge_cells('F2:F3')
    #     ws.column_dimensions['G'].width = 20
    #     ws['G2'] = 'Estado conservacion'
    #     ws.merge_cells('G2:G3')
    #     ws.column_dimensions['H'].width = 20
    #     ws['H2'] = 'Domicilio'
    #     ws.merge_cells('H2:H3')
    #
    #     self.estilosCelda(ws['I1'])
    #     ws['I1'] = 'Implantacion y Comprtamiento Resolucion del SIEN'
    #     ws.merge_cells('I1:J1')
    #     ws.column_dimensions['I'].width = 20
    #     ws['I2'] = 'Convenio'
    #     ws.merge_cells('I2:I3')
    #     ws.column_dimensions['J'].width = 20
    #     ws['J2'] = 'Firmado Director'
    #     ws.merge_cells('J2:J3')
    #
    #     self.estilosCelda(ws['K1'])
    #     ws['K1'] = 'Disciplina informativa acumulada hasta el cierre del mes anterior (SIEN)'
    #     ws.merge_cells('K1:N1')
    #     ws.column_dimensions['K'].width = 20
    #     ws['K2'] = 'Total de modelos a  reportar'
    #     ws.merge_cells('K2:K3')
    #     ws.column_dimensions['L'].width = 20
    #     ws['L2'] = 'Reportados fuera fecha'
    #     ws.merge_cells('L2:L3')
    #     ws.column_dimensions['M'].width = 20
    #     ws['M2'] = 'Reportados en fecha'
    #     ws.merge_cells('M2:M3')
    #     ws.column_dimensions['N'].width = 20
    #     ws['N2'] = 'No reportados'
    #     ws.merge_cells('N2:N3')
    #
    #     self.estilosCelda(ws['O1'])
    #     ws['O1'] = 'Calidad de la informacion'
    #     ws.merge_cells('O1:P1')
    #     ws.column_dimensions['O'].width = 20
    #     ws['O2'] = 'Señalamiento de errores'
    #     ws.merge_cells('O2:O3')
    #     ws.column_dimensions['P'].width = 20
    #     ws['P2'] = 'Cantidad de señalamientos'
    #     ws.merge_cells('P2:P3')
    #
    #     self.estilosCelda(ws['Q1'])
    #     ws['Q1'] = 'Asesoramiento metodologiaco'
    #     ws.merge_cells('Q1:S1')
    #     ws.column_dimensions['Q'].width = 20
    #     ws['Q2'] = 'Resivio asesoramiento'
    #     ws.merge_cells('Q2:Q3')
    #     ws.column_dimensions['R'].width = 20
    #     ws['R2'] = 'Posee bases metodologicas del SIEN'
    #     ws.merge_cells('R2:R3')
    #     ws.column_dimensions['S'].width = 20
    #     ws['S2'] = 'Tipo de soporte'
    #     ws.merge_cells('S2:S3')
    #
    #     self.estilosCelda(ws['T1'])
    #     ws['T1'] = 'Cobertura'
    #     ws.merge_cells('T1:V1')
    #     ws.column_dimensions['T'].width = 20
    #     ws['T2'] = 'Establecimientos asociados'
    #     ws.merge_cells('T2:T3')
    #     ws.column_dimensions['U'].width = 20
    #     ws['U2'] = 'Cuantos'
    #     ws.merge_cells('U2:U3')
    #     ws.column_dimensions['V'].width = 20
    #     ws['V2'] = 'Con contabilidad propia'
    #     ws.merge_cells('V2:V3')
    #
    #     self.estilosCelda(ws['W1'])
    #     ws['W1'] = 'Atencion a las estadisticas'
    #     ws.merge_cells('W1:Z1')
    #     ws.column_dimensions['W'].width = 20
    #     ws['W2'] = 'Estructura para atender la actv. estadistica'
    #     ws.merge_cells('W2:W3')
    #     ws.column_dimensions['X'].width = 20
    #     ws['X2'] = 'Posee personal capacitado para brindar informacion estadistica'
    #     ws.merge_cells('X2:X3')
    #     ws.column_dimensions['Y'].width = 20
    #     ws['Y2'] = 'Esta incluido en el Plan de Prevencion del centro como un punto vulnerable la infomacion estadistica'
    #     ws.merge_cells('Y2:Y3')
    #     ws.column_dimensions['Z'].width = 20
    #     ws['Z2'] = 'Utiliza la infomacion estadistica para la toma de decisiones'
    #     ws.merge_cells('Z2:Z3')
    #     ws['A4'].fill = PatternFill(start_color="92a2ab", end_color="92a2ab", fill_type="solid")
    #     ws.merge_cells('A4:Z4')
    #
    #     cont = 5
    #
    #     for cuestionario in query_cuestionario:
    #         ws.cell(row=cont, column=1).value = cuestionario.entidad_codigo.nombre_CI
    #         ws.cell(row=cont, column=2).value = cuestionario.entidad_codigo.codigo_CI
    #         ws.cell(row=cont, column=3).value = str(cuestionario.entidad_codigo.ome_codigo)
    #         preguntas = self.getPreguntasDelCuestionario(cuestionario.id)
    #         self.pintarDatos(ws,cont,preguntas)
    #
    #         cont += 1
    #
    #     nombre_archivo = "ReporteGeneral.xls"
    #     response = HttpResponse(content_type="aplication/ms-excel")
    #     content = "attachment; filename = {0}".format(nombre_archivo)
    #     response['Content-Disposition'] = content
    #     wb.save(response)
    #     return response
    #
    # def getCuestionarios(self):
    #    if self.request.user.is_superuser:
    #         query = cuestionario.objects.all()
    #         return query
    #    elif self.request.user.has_perm('guiaEstadistica.pinar'):
    #         query = cuestionario.objects.filter(entidad_codigo__ote_codigo=21)
    #         return query
    #    elif self.request.user.has_perm('guiaEstadistica.habana'):
    #        query = cuestionario.objects.filter(entidad_codigo__ote_codigo=23)
    #        return query
    #
    #
    # def listadoPreguntas(self, cuestionario):
    #     query = PreguntasEvaluadas.objects.filter(captacion_id__id=cuestionario)
    #     return query
    #
    # def getPreguntasDelCuestionario(self, cuestionario):
    #     query = PreguntasEvaluadas.objects.filter(captacion_id__id=cuestionario)
    #     return query
    #
    # def estilosCelda(self,celda):
    #     celda.alignment = Alignment(horizontal='center', vertical='center')
    #     celda.border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
    #                           bottom=Side(border_style="thin"))
    #     celda.fill = PatternFill(start_color="66FFCC", end_color="66FFCC", fill_type="solid")
    #     celda.font = Font(name='Broadway', size=12, bold=True)
    #
    # def pintarDatos(self, ws, cont, preguntasDelCuestionario):
    #     columna = 4
    #     aux = preguntasDelCuestionario[::-1]
    #     for p in aux[4:]:
    #         ws.cell(row=cont, column=columna).value = p.respuesta
    #         columna += 1

    template_name = 'reportes/general.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte General'
        context['grupoPreguntas'] = self.getGrupoPreguntas()
        context['cuestionarios'] = self.getCuestionarios()
        return context

    def getGrupoPreguntas(self):
        data = {}
        query = clasificadorIndicadores.objects.filter(seccion_id__guia_id__activo=True)
        for i in query[1:]:
            data[i.id] = i.nombre
        return data

    def getCuestionarios(self):
        if self.request.user.is_superuser:
            query = cuestionario.objects.all()
            return query
        elif self.request.user.has_perm('guiaEstadistica.pinar'):
            query = cuestionario.objects.filter(entidad_codigo__ote_codigo=21,guia__activo=True)
            return query
        elif self.request.user.has_perm('guiaEstadistica.habana'):
            query = cuestionario.objects.filter(entidad_codigo__ote_codigo=23,guia__activo=True)
            return query

class crearGuiaDefinida(TemplateView):

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        activo = None
        query = guiaEstadistica.objects.all()
        action = request.POST['action']
        if request.POST['activo'] == 'true':
            activo = True
            for i in query:
                i.activo = False
                i.save()
        else:
            activo = False
        try:
            if action == 'creacionDeGuia':
                guiaNueva=guiaEstadistica(
                    nombre=request.POST['guiaNueva'],
                    activo=activo
                )
                guiaNueva.save()
                query_secciones = seccion.objects.filter(guia_id__nombre=request.POST['guiaYaDefinida'])
                for i in query_secciones:
                    if i.tipo == 1 or i.tipo == 2:
                        aux = seccion(
                            nombre=i.nombre,
                            guia_id=guiaNueva,
                            periodo_id=None,
                            numero=None,
                            subNumero=None,
                            orden=i.orden,
                            activo=i.activo,
                            tipo=i.tipo
                        )
                        aux.save()
                        self.crearGrupoPreguntas(i, guiaNueva)
                    else:
                       aux=seccion(
                            nombre=i.nombre,
                            guia_id=guiaEstadistica.objects.get(id=guiaNueva.id),
                            periodo_id=clasificadorPeriodo.objects.get(id=i.periodo_id.id),
                            numero=i.numero,
                            subNumero=i.subNumero,
                            orden=i.orden,
                            activo=i.activo,
                            tipo=i.tipo
                        )
                       aux.save()
                data['exito'] = 'La guia '+ guiaNueva.nombre + 'ha sido creada correctamente.'
            else:
                data['error'] = 'error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

    def crearGrupoPreguntas(self, objSeccion, guia):
        query = seccion.objects.filter(guia_id__id=guia.id)
        for i in query:
            if i.nombre == objSeccion.nombre:
                grupoPreguntas = clasificadorIndicadores.objects.filter(seccion_id__id=objSeccion.id)
                for j in grupoPreguntas:
                    aux = clasificadorIndicadores(
                        seccion_id=seccion.objects.get(id=i.id),
                        nombre=j.nombre,
                        activo=True
                    )
                    aux.save()
                    self.crearPreguntas(j, aux)

    def crearPreguntas(self, grupoPreguntaGuiaDefinida, grupoPreguntaGuiaNueva):
        query = Indicadores.objects.filter(clasificadorIndicadores_id__id=grupoPreguntaGuiaDefinida.id)
        if grupoPreguntaGuiaDefinida.nombre == grupoPreguntaGuiaNueva.nombre:
            for i in query:
                if i.cod_indicador == None:
                    try:
                        aux = Indicadores(
                            clasificadorIndicadores_id=clasificadorIndicadores.objects.get(id=grupoPreguntaGuiaNueva.id),
                            nombre=i.nombre,
                        )
                        aux.save()
                        self.asignarRespuesta(aux, i.respuestas_id.all())
                    except:
                        print('error')
                else:
                    try:
                        aux = Indicadores(
                            clasificadorIndicadores_id=clasificadorIndicadores.objects.get(id=grupoPreguntaGuiaNueva.id),
                            nombre=i.nombre,
                            cod_indicador=i.cod_indicador
                        )
                        aux.save()
                        self.asignarRespuesta(aux, i.respuestas_id.all())
                    except:
                        print('error')

    def asignarRespuesta(self, preguntaNueva, listaRespuestasDefinidas):
        for i in listaRespuestasDefinidas:
            preguntaNueva.respuestas_id.add(i)

class reporteVerificacionIndicadores(TemplateView):
    template_name = 'reportes/verificacion.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte de Verificacion de Indicadores'
        context['secciones'] = self.getSecciones()
        return context

    def getSecciones(self):
        data = {}
        secciones = seccion.objects.filter(guia_id__activo=True)
        for i in secciones:
            verificados = 0
            if i.numero != None:
                query = verificacion.objects.filter(seccion_id__id=i.id)
                if query.count()!= 0:
                    for j in query:
                        verificados += j.indicadoresVerificados
                    data[i.nombre]=verificados
        return data

class reporteDisciplinaInformativa(TemplateView):
    template_name = 'reportes/disciplinaInformativa.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte de Disciplina Informativa'
        context['cuestionarios'] = self.listaCuestionario()
        return context

    def listaCuestionario(self):
        query = getCuestionarios(self.request.user)
        return query

class reporteSeñalamientosErrores(TemplateView):
    template_name = 'reportes/señalamientosErrores.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte de Señalamientos Errores'
        return context

class reporteDisciplinaInformativaCentroControlado(TemplateView):
    template_name = 'reportes/disciplinaInfoCentrosControlados.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte de Disciplina Informativa Centros Controlados'
        return context

class reporteDomicilioSocial(TemplateView):
    template_name = 'reportes/domicilioSocial.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte de Domicilio Social Incorrecto'
        return context

class reporteDeficiencias(TemplateView):
    template_name = 'reportes/deficiencias.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titulo'] = 'Reporte de Deficiencias'
        return context
